import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def gerar_excel_orcamento(nome_cliente, endereco_obra, data_orcamento, itens, numero_orcamento, forma_pagamento):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento"

    header_font = Font(bold=True, size=11)
    red_fill = PatternFill("solid", fgColor="CC0000")
    grey_fill = PatternFill("solid", fgColor="DDDDDD")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "MARMORARIA EBENEZER MARMORES E GRANITOS"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")

    ws["A2"] = "RUA SAFIRA, 842 - INDAIATUBA/SP | (19) 99514-0294 | ebenezermarmoraria2020@gmail.com"
    ws.merge_cells("A2:I2")

    ws["A4"] = "CLIENTE:"
    ws["B4"] = nome_cliente
    ws["A4"].font = Font(bold=True)

    ws["A5"] = "ENDEREÇO:"
    ws["B5"] = endereco_obra

    ws["G4"] = "DATA:"
    ws["H4"] = data_orcamento
    ws["G4"].font = Font(bold=True)

    ws["G5"] = "ORÇAMENTO Nº:"
    try:
        ws["H5"] = f"{int(numero_orcamento):03d}"
    except Exception:
        ws["H5"] = str(numero_orcamento)
    ws["G5"].font = Font(bold=True)

    headers = ["AMBIENTE", "DESCRIÇÃO", "MATERIAL", "QTD", "ML", "PREÇO/ML", "ÁREA m²", "VALOR/m²", "TOTAL"]
    col_widths = [18, 30, 22, 8, 10, 14, 12, 14, 16]

    row = 7
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = red_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    from collections import OrderedDict
    grupos = OrderedDict()
    for item in itens:
        amb = item.get("ambiente", "")
        if amb not in grupos:
            grupos[amb] = []
        grupos[amb].append(item)

    row = 8
    total_geral = 0.0

    for nome_amb, grupo in grupos.items():
        subtotal = 0.0
        for item in grupo:
            ml = float(item.get("ml", 0) or 0)
            area = float(item.get("area", 0) or 0)
            valor_m2 = float(item.get("valor_m2", 0) or 0)
            preco_ml = float(item.get("preco_ml", 0) or 0)
            total_item = float(item.get("total", 0) or 0)

            values = [
                nome_amb,
                item.get("descricao", ""),
                item.get("material", "-"),
                float(item.get("qtd", 1) or 1),
                ml if ml > 0 else None,
                preco_ml if preco_ml > 0 else None,
                area if area > 0 else None,
                valor_m2 if valor_m2 > 0 else None,
                total_item,
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = border
                cell.fill = white_fill
                if col_idx >= 4:
                    cell.alignment = center
                    if col_idx in (6, 8, 9) and val is not None:
                        cell.number_format = 'R$ #,##0.00'

            subtotal += total_item
            total_geral += total_item
            row += 1

        sub_cell = ws.cell(row=row, column=1, value=f"SUBTOTAL {nome_amb}")
        sub_cell.font = Font(bold=True)
        sub_cell.fill = grey_fill
        sub_cell.border = border
        ws.merge_cells(f"A{row}:H{row}")
        total_cell = ws.cell(row=row, column=9, value=subtotal)
        total_cell.font = Font(bold=True)
        total_cell.fill = grey_fill
        total_cell.border = border
        total_cell.number_format = 'R$ #,##0.00'
        row += 1

    row += 1
    total_label = ws.cell(row=row, column=8, value="TOTAL GERAL:")
    total_label.font = Font(bold=True, size=12)
    total_value = ws.cell(row=row, column=9, value=total_geral)
    total_value.font = Font(bold=True, size=12)
    total_value.number_format = 'R$ #,##0.00'

    row += 1
    enc_label = ws.cell(row=row, column=8, value="COM ENCARGOS (8%):")
    enc_label.font = Font(bold=True, color="CC0000")
    enc_value = ws.cell(row=row, column=9, value=total_geral * 1.08)
    enc_value.font = Font(bold=True, color="CC0000")
    enc_value.number_format = 'R$ #,##0.00'

    row += 2
    ws.cell(row=row, column=1, value=f"Forma de pagamento: {forma_pagamento}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def gerar_excel_montador(nome_cliente, data_orcamento, itens, total_geral):
    wb = Workbook()
    ws = wb.active
    ws.title = "MONTADOR"

    ws["A1"] = "MARMORARIA EBENEZER MARMORES E GRANITOS"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:F1")

    ws["A3"] = "Cliente:"
    ws["B3"] = nome_cliente
    ws["A3"].font = Font(bold=True)

    ws["A4"] = "Data:"
    ws["B4"] = data_orcamento

    ws["A6"] = "SERVIÇO DE MONTADOR TERCEIRIZADO"
    ws["A6"].font = Font(bold=True, size=11)
    ws.merge_cells("A6:F6")

    headers = ["DESCRIÇÃO", "QTD", "ML", "VALOR/ML", "VALOR FIXO", "TOTAL"]
    col_widths = [35, 8, 10, 14, 14, 14]

    red_fill = PatternFill("solid", fgColor="CC0000")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=8, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = red_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row = 9
    for item in itens:
        vals = [
            item.get("descricao", ""),
            float(item.get("qtd", 1) or 1),
            float(item.get("ml", 0) or 0),
            float(item.get("valor_ml", 0) or 0),
            float(item.get("valor_fixo", 0) or 0),
            float(item.get("total", 0) or 0),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
            if col_idx >= 4:
                cell.number_format = 'R$ #,##0.00'
        row += 1

    row += 1
    ws.cell(row=row, column=5, value="TOTAL:").font = Font(bold=True)
    total_cell = ws.cell(row=row, column=6, value=total_geral)
    total_cell.font = Font(bold=True)
    total_cell.number_format = 'R$ #,##0.00'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
